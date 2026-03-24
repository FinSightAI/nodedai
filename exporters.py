"""
Export price history to Excel / CSV.
"""
import json
import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Optional, List

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

import database as db


# ── Excel ──────────────────────────────────────────────────────────────────────

def export_excel(watch_ids: Optional[List[int]] = None) -> bytes:
    """
    Export price history to a styled Excel file.
    Returns bytes to stream to user.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    items = db.get_all_watch_items(enabled_only=False)
    if watch_ids:
        items = [i for i in items if i["id"] in watch_ids]

    # ── Style helpers ──────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
    HEADER_FONT = Font(color="00d4ff", bold=True, size=11)
    SUBHEADER_FILL = PatternFill("solid", fgColor="16213e")
    SUBHEADER_FONT = Font(color="ffffff", bold=True)
    GOOD_FILL = PatternFill("solid", fgColor="1a472a")
    BAD_FILL  = PatternFill("solid", fgColor="4a1942")
    CENTER = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="333333")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("📊 סיכום")
    ws_sum.sheet_view.rightToLeft = True
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 12
    ws_sum.column_dimensions["E"].width = 12
    ws_sum.column_dimensions["F"].width = 15

    # Title
    ws_sum["A1"] = "✈️ MegaTraveller — דוח מחירים"
    ws_sum["A1"].font = Font(size=16, bold=True, color="00d4ff")
    ws_sum["A2"] = f"נוצר: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum["A2"].font = Font(color="888888", italic=True)

    headers = ["פריט", "קטגוריה", "יעד", "מחיר נוכחי", "מינימום", "בדיקות"]
    for col, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    row = 5
    for item in items:
        history = db.get_price_history(item["id"], limit=200)
        last = db.get_last_price(item["id"])
        lowest = db.get_lowest_price(item["id"])

        prices_row = [
            item["name"],
            item["category"],
            item["destination"],
            f"{last['price']:.0f} {last['currency']}" if last else "—",
            f"{lowest['price']:.0f}" if lowest else "—",
            len(history),
        ]

        for col, val in enumerate(prices_row, 1):
            cell = ws_sum.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            if col == 4 and last and lowest and last["price"] <= lowest["price"] * 1.05:
                cell.fill = GOOD_FILL
                cell.font = Font(color="00ff88", bold=True)

        row += 1

    # ── Per-item sheets ────────────────────────────────────────────────────────
    for item in items:
        history = db.get_price_history(item["id"], limit=200)
        if not history:
            continue

        safe_name = item["name"][:28].replace("/", "-").replace(":", "-")
        ws = wb.create_sheet(safe_name)
        ws.sheet_view.rightToLeft = True

        # Header
        ws["A1"] = item["name"]
        ws["A1"].font = Font(size=14, bold=True, color="00d4ff")
        ws["A2"] = f"{item['category']} | {item['destination']}"
        ws["A2"].font = Font(color="888888")

        col_headers = ["תאריך", "מחיר", "מטבע", "מקור", "פרטים"]
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 35

        for col, h in enumerate(col_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

        prices = [r["price"] for r in history]
        min_p = min(prices)
        max_p = max(prices)

        history_rev = list(reversed(history))
        for r_idx, record in enumerate(history_rev, 5):
            price = record["price"]
            details_obj = {}
            try:
                details_obj = json.loads(record.get("details", "{}"))
            except Exception:
                pass

            row_data = [
                record["checked_at"][:16].replace("T", " "),
                price,
                record["currency"],
                record["source"][:20],
                details_obj.get("details", "")[:50],
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.border = BORDER
                cell.alignment = CENTER

            # Color code price column
            price_cell = ws.cell(row=r_idx, column=2)
            if price == min_p:
                price_cell.fill = GOOD_FILL
                price_cell.font = Font(color="00ff88", bold=True)
            elif price == max_p:
                price_cell.fill = BAD_FILL
                price_cell.font = Font(color="ff4444", bold=True)

        # ── Line chart ─────────────────────────────────────────────────────────
        if len(history_rev) >= 2:
            chart = LineChart()
            chart.title = f"מחירים — {item['name']}"
            chart.style = 10
            chart.y_axis.title = "מחיר"
            chart.x_axis.title = "תאריך"

            price_data = Reference(
                ws,
                min_col=2, max_col=2,
                min_row=4, max_row=4 + len(history_rev)
            )
            chart.add_data(price_data, titles_from_data=True)
            chart.shape = 4

            last_row = 4 + len(history_rev)
            ws.add_chart(chart, f"G4")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── CSV ────────────────────────────────────────────────────────────────────────

def export_csv(watch_id: int) -> str:
    """Export single item history as CSV string."""
    history = db.get_price_history(watch_id, limit=1000)
    item_rows = db.get_all_watch_items(enabled_only=False)
    item = next((i for i in item_rows if i["id"] == watch_id), {})

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["# MegaTraveller Price Export"])
    writer.writerow([f"# Item: {item.get('name', '')}"])
    writer.writerow([f"# Exported: {datetime.now().isoformat()}"])
    writer.writerow([])
    writer.writerow(["Date", "Price", "Currency", "Source", "Details"])

    for r in reversed(history):
        details_obj = {}
        try:
            details_obj = json.loads(r.get("details", "{}"))
        except Exception:
            pass
        writer.writerow([
            r["checked_at"][:16].replace("T", " "),
            r["price"],
            r["currency"],
            r["source"],
            details_obj.get("details", ""),
        ])

    return output.getvalue()
